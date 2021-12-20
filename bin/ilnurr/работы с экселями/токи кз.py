import openpyxl
from openpyxl import load_workbook
import pandas as pd
import numpy as np

wb = load_workbook('чтобы смотреть.xlsx')
ws = wb['ИЛ-1 - ИЛ-2 (13) мин.']

i = 0
dfs = pd.read_excel(io='чтобы смотреть.xlsx', engine='openpyxl', sheet_name=['ИЛ-1 - ИЛ-2 (13) мин.'])
dfs = pd.concat(dfs).reset_index(drop=True)
print(dfs)

Node1_list = list(dfs['узел 1'])
Node2_list = list(dfs['узел 2'])
z_list = list(dfs['z'])

checked_List = list(dfs['checked'])
node_List = list(dfs['узел'])
u_List = list(dfs['u'])
Ik_List = list(dfs['ik'])
zc_List = list(dfs['zc'])

print(Node1_list)

values1 = np.array(Node1_list)
values2 = np.array(Node2_list)
startNode = ['172']
while i < len(startNode):
    findedNodes1 = np.where(values1 == startNode[i])[0]
    findedNodes2 = np.where(values2 == startNode[i])[0]
    Zc1 = float(u_List[node_List.index(startNode[i])]) / ((3 ** 0.5) * float(Ik_List[node_List.index(startNode[i])]))
    print('U', u_List[node_List.index(startNode[i])], 'Ik', Ik_List[node_List.index(startNode[i])])
    ws['K' + str(node_List.index(startNode[i]) + 2)].value = Zc1
    Ik1 = float(Ik_List[node_List.index(startNode[i])])
    Uc1 = float(u_List[node_List.index(startNode[i])])
    print(startNode[i])
    for node1 in range(len(findedNodes1)):
        try:
            if checked_List[node_List.index(Node2_list[findedNodes1[node1]])] != 1:
                checked_List[node_List.index(Node2_list[findedNodes1[node1]])] = 1
                startNode.append(node_List[node_List.index(Node2_list[findedNodes1[node1]])])
                Z = float(z_list[findedNodes1[node1]])
                Ik2 = Uc1 / ((3 ** 0.5) * (Zc1 + Z))
                Ik_List[node_List.index(Node2_list[findedNodes1[node1]])] = Ik2
                print(findedNodes1[node1], node_List.index(Node2_list[findedNodes1[node1]]))
                ws['J' + str(node_List.index(Node2_list[findedNodes1[node1]]) + 2)].value = Ik2
        except:
            continue
    for node2 in range(len(findedNodes2)):
        try:
            if checked_List[node_List.index(Node1_list[findedNodes2[node2]])] != 1:
                checked_List[node_List.index(Node1_list[findedNodes2[node2]])] = 1
                startNode.append(node_List[node_List.index(Node1_list[findedNodes2[node2]])])
                Z = float(z_list[findedNodes2[node2]])
                Ik2 = Uc1 / ((3 ** 0.5) * (Zc1 + Z))
                Ik_List[node_List.index(Node1_list[findedNodes2[node2]])] = Ik2
                print(findedNodes2[node2], node_List.index(Node1_list[findedNodes2[node2]]))
                ws['J' + str(node_List.index(Node1_list[findedNodes2[node2]]) + 2)].value = Ik2
        except:
            continue

    i += 1
wb.save('Обработка.xlsx')
