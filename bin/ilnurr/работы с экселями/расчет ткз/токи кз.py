import openpyxl
# import warnings
# warnings.simplefilter(action='ignore', category=FutureWarning)
from openpyxl import load_workbook
import pandas as pd
import numpy as np

wb = load_workbook('чтобы смотреть.xlsx')

zzzz = -1
for sheet in wb.worksheets:
    zzzz += 1
    if zzzz < 4:
        continue

    sheetName = str(sheet).split('"')[1]
    # if sheetName != 'НД-67 МИН.' and sheetName != 'НД-67 МАКС.':
    #     print(sheet)
    #     continue
    print(sheet)
    ws = wb[sheetName]
    print('\n\n' + sheetName + '\n\n')
    i = 0
    dfs = pd.read_excel(io='чтобы смотреть.xlsx', engine='openpyxl', sheet_name=[sheetName])
    dfs = pd.concat(dfs).reset_index(drop=True)
    print(dfs)

    Node1_list = list(dfs['узел 1'])
    Node2_list = list(dfs['узел 2'])
    z_list = list(dfs['z'])

    checked_List = list(dfs['checked'])
    node_List = list(dfs['узел'])
    u_List = list(dfs['u'])
    Ik_List = list(dfs['ik'])
    zc_List = list(dfs['zc'])

    print(Node1_list)

    values1 = np.array(Node1_list)
    values2 = np.array(Node2_list)

    startIndex = next(x for x in range(len(checked_List)) if checked_List[x] == 1.0)
    startNode = [node_List[startIndex]]
    print('\n', startNode, startIndex, '\n')

    while i < len(startNode):
        findedNodes1 = np.where(values1 == startNode[i])[0]
        findedNodes2 = np.where(values2 == startNode[i])[0]
        Zc1 = float(u_List[node_List.index(startNode[i])]) / ((3 ** 0.5) * float(Ik_List[int(node_List.index(startNode[i]))]))
        print('U', u_List[node_List.index(startNode[i])], 'Ik', Ik_List[node_List.index(startNode[i])])
        ws['K' + str(node_List.index(startNode[i]) + 2)].value = Zc1
        Ik1 = float(Ik_List[node_List.index(startNode[i])])
        Uc1 = float(u_List[node_List.index(startNode[i])])
        print(startNode[i])
        for node1 in range(len(findedNodes1)):
            try:
                if checked_List[node_List.index(Node2_list[findedNodes1[node1]])] != 1:
                    checked_List[node_List.index(Node2_list[findedNodes1[node1]])] = 1
                    startNode.append(node_List[node_List.index(Node2_list[findedNodes1[node1]])])
                    Z = float(z_list[findedNodes1[node1]])
                    Ik2 = Uc1 / ((3 ** 0.5) * (Zc1 + Z))
                    Ik_List[node_List.index(Node2_list[findedNodes1[node1]])] = Ik2
                    print(findedNodes1[node1], node_List.index(Node2_list[findedNodes1[node1]]))
                    ws['J' + str(node_List.index(Node2_list[findedNodes1[node1]]) + 2)].value = Ik2
            except:
                continue
        for node2 in range(len(findedNodes2)):
            try:
                if checked_List[node_List.index(Node1_list[findedNodes2[node2]])] != 1:
                    checked_List[node_List.index(Node1_list[findedNodes2[node2]])] = 1
                    startNode.append(node_List[node_List.index(Node1_list[findedNodes2[node2]])])
                    Z = float(z_list[findedNodes2[node2]])
                    Ik2 = Uc1 / ((3 ** 0.5) * (Zc1 + Z))
                    Ik_List[node_List.index(Node1_list[findedNodes2[node2]])] = Ik2
                    print(findedNodes2[node2], node_List.index(Node1_list[findedNodes2[node2]]))
                    ws['J' + str(node_List.index(Node1_list[findedNodes2[node2]]) + 2)].value = Ik2
            except:
                continue

        i += 1

wb.save('Обработка.xlsx')
